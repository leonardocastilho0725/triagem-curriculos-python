import os
import re
import pdfplumber
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# --- REGRAS DE NEGÓCIO ---
IDADE_MINIMA = 18
IDADE_MAXIMA = 23
ANO_ATUAL = 2026  # Ano base do sistema atual

# Cidades mapeadas com todas as suas variações comuns de digitação
MAPA_CIDADES = {
    "Itaquaquecetuba": ["itaquaquecetuba", "itaqua", "itaquá"],
    "Poá": ["poá", "poa"],
    "Mogi das Cruzes": ["mogi das cruzes", "mogi", "mogy"],
    "Ferraz de Vasconcelos": ["ferraz de vasconcelos", "ferraz"],
    "Suzano": ["suzano"]
}


def extrair_texto_pdf(caminho_pdf):
    texto_completo = ""
    try:
        with pdfplumber.open(caminho_pdf) as pdf:
            for pagina in pdf.pages:
                texto_completo += pagina.extract_text() or ""
    except Exception as e:
        print(f"Erro ao ler {caminho_pdf}: {e}")
    return texto_completo


def extrair_dados_curriculo(texto):
    texto_minusculo = texto.lower()

    # 1. CAPTURA DE E-MAIL
    email = "Não encontrado"
    padrao_email = re.search(r'[\w\.-]+@[\w\.-]+\.\w+', texto_minusculo)
    if padrao_email:
        email = padrao_email.group(0)

    # 2. CAPTURA DE TELEFONE / WHATSAPP
    # Agora usamos finditer e damos preferência ao número com mais dígitos
    # e formato de telefone plausível, em vez de aceitar cegamente o primeiro
    # match (que podia ser um CPF, CEP ou data).
    telefone = "Não encontrado"
    padrao_tel_completo = re.compile(
        r'(?:\+?55\s?)?\(?\d{2}\)?[\s.-]?9?\d{4}[-.\s]?\d{4}'
    )
    candidatos = []
    for m in padrao_tel_completo.finditer(texto_minusculo):
        trecho = m.group(0)
        digitos = re.sub(r'\D', '', trecho)
        # Telefone BR: 10 ou 11 dígitos (com DDD), com ou sem +55 (12/13 dígitos)
        if len(digitos) in (10, 11, 12, 13):
            candidatos.append((len(digitos), trecho.strip()))
    if candidatos:
        # Prioriza o candidato com mais dígitos (mais completo/plausível)
        candidatos.sort(key=lambda x: x[0], reverse=True)
        telefone = candidatos[0][1]

    # 3. VALIDAÇÃO DE IDADE
    # Ordem de confiabilidade: data de nascimento completa > "Idade: X" >
    # ano isolado > "X anos" (o mais ambíguo, pode ser "5 anos de experiência",
    # por isso vai por último e só é usado se nada mais funcionar).
    idade = None

    padrao_data_nasc = re.search(r'\b(\d{2})[\/\.-](\d{2})[\/\.-](\d{4})\b', texto_minusculo)
    padrao_idade_direta = re.search(r'idade[\s:]*(\d{2})', texto_minusculo)
    padrao_ano_isolado = re.search(r'\b(19\d{2}|20\d{2})\b', texto_minusculo)
    padrao_idade_generico = re.search(r'(\d{2})\s*anos', texto_minusculo)

    if padrao_data_nasc:
        ano_nascimento = int(padrao_data_nasc.group(3))
        idade = ANO_ATUAL - ano_nascimento
    elif padrao_idade_direta:
        idade = int(padrao_idade_direta.group(1))
    elif padrao_ano_isolado:
        ano_detectado = int(padrao_ano_isolado.group(1))
        if ano_detectado < ANO_ATUAL - 10:
            idade = ANO_ATUAL - ano_detectado
    elif padrao_idade_generico:
        idade = int(padrao_idade_generico.group(1))

    # Sanidade: descarta valores fora de uma faixa humana plausível
    if idade is not None and not (14 <= idade <= 90):
        idade = None

    # 4. VALIDAÇÃO DE CIDADE FLEXÍVEL
    cidade_encontrada = "Não identificada"
    for cidade_oficial, variacoes in MAPA_CIDADES.items():
        for var in variacoes:
            if re.search(r'\b' + re.escape(var) + r'\b', texto_minusculo):
                cidade_encontrada = cidade_oficial
                break
        if cidade_encontrada != "Não identificada":
            break

    # 5. VALIDAÇÃO DE ESCOLARIDADE
    # Removida a heurística de "texto curto = ensino médio completo": ela gerava
    # falsos positivos para currículos curtos de quem tem o ensino médio
    # incompleto. Agora só marcamos como completo quando há evidência textual real.
    termos_ensino_medio = ["ensino médio", "ensino medio", "segundo grau", "2º grau", "2o grau", "colegial"]
    termos_conclusao = ["completo", "concluído", "concluido", "formado", "graduado", "conclusao"]
    termos_incompleto = ["incompleto", "cursando o ensino médio", "cursando o ensino medio"]
    termos_superior = ["superior", "faculdade", "graduação", "graduacao", "universidade", "tecnólogo", "tecnologo"]

    ensino_medio_completo = False

    tem_incompleto = any(termo in texto_minusculo for termo in termos_incompleto)

    if not tem_incompleto and any(sup in texto_minusculo for sup in termos_superior):
        # Cursar/ter cursado ensino superior pressupõe ensino médio completo
        ensino_medio_completo = True
    elif any(termo in texto_minusculo for termo in termos_ensino_medio):
        if any(concl in texto_minusculo for concl in termos_conclusao) and not tem_incompleto:
            ensino_medio_completo = True

    return idade, cidade_encontrada, ensino_medio_completo, email, telefone


def triagem_curriculo(caminho_pdf):
    texto = extrair_texto_pdf(caminho_pdf)
    idade, cidade, ensino_medio_completo, email, telefone = extrair_dados_curriculo(texto)

    cidade_valida = cidade != "Não identificada"
    idade_valida = idade is not None and (IDADE_MINIMA <= idade <= IDADE_MAXIMA)
    escolaridade_valida = ensino_medio_completo

    if cidade_valida and idade_valida and escolaridade_valida:
        status = "APROVADO"
        motivo = "Atende todos os critérios exigidos."
    else:
        status = "REPROVADO"
        motivos = []
        if not idade_valida:
            motivos.append(f"Idade fora da faixa (Detectado: {idade if idade else 'N/I'})")
        if not cidade_valida:
            motivos.append(f"Cidade fora do eixo ampliado (Detectado: {cidade})")
        if not escolaridade_valida:
            motivos.append("Ensino Médio incompleto/não identificado")
        motivo = " | ".join(motivos)

    return {
        "Nome do Arquivo": os.path.basename(caminho_pdf),
        "Telefone/WhatsApp": telefone,
        "E-mail": email,
        "Idade": idade if idade else "Não Encontrada",
        "Cidade": cidade,
        "Escolaridade": "Ensino Médio OK" if escolaridade_valida else "Incompleto/N/I",
        "Status": status,
        "Motivo Detalhado": motivo
    }


# --- PROCESSAMENTO PRINCIPAL ---
pasta_curriculos = "./curriculos"

if not os.path.exists(pasta_curriculos):
    os.makedirs(pasta_curriculos)
    print(f"Pasta '{pasta_curriculos}' pronta. Adicione os PDFs nela.")
else:
    lista_resultados = []
    print("⏳ Efetuando triagem corporativa inteligente e robusta...")
    for arquivo in os.listdir(pasta_curriculos):
        if arquivo.endswith(".pdf"):
            caminho_completo = os.path.join(pasta_curriculos, arquivo)
            resultado = triagem_curriculo(caminho_completo)
            lista_resultados.append(resultado)

    if lista_resultados:
        df_geral = pd.DataFrame(lista_resultados)
        df_geral = df_geral.sort_values(by="Nome do Arquivo")

        df_aprovados = df_geral[df_geral["Status"] == "APROVADO"]
        df_reprovados = df_geral[df_geral["Status"] == "REPROVADO"]

        nome_planilha = "resultado_triagem.xlsx"

        with pd.ExcelWriter(nome_planilha, engine="openpyxl") as writer:
            df_aprovados.to_excel(writer, sheet_name="Aprovados", index=False)
            df_reprovados.to_excel(writer, sheet_name="Reprovados", index=False)

            workbook = writer.book

            fonte_cabecalho = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            fonte_dados = Font(name="Calibri", size=11, color="000000")
            fonte_aprovado = Font(name="Calibri", size=11, bold=True, color="276A3C")
            fonte_reprovado = Font(name="Calibri", size=11, bold=True, color="9C0006")

            fill_cabecalho = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            fill_zebra = PatternFill(start_color="F2F4F7", end_color="F2F4F7", fill_type="solid")
            fill_status_aprovado = PatternFill(start_color="D5EAD8", end_color="D5EAD8", fill_type="solid")
            fill_status_reprovado = PatternFill(start_color="FFD6D6", end_color="FFD6D6", fill_type="solid")

            borda_fina = Border(
                left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
                top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
            )

            for nome_aba in ["Aprovados", "Reprovados"]:
                if nome_aba not in workbook.sheetnames:
                    continue

                worksheet = workbook[nome_aba]

                # Corrigido: row_dimensions é indexado por número de linha,
                # não tem um atributo .height direto.
                worksheet.row_dimensions[1].height = 26

                for col_num in range(1, worksheet.max_column + 1):
                    cell = worksheet.cell(row=1, column=col_num)
                    cell.fill = fill_cabecalho
                    cell.font = fonte_cabecalho
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = borda_fina

                for row_num in range(2, worksheet.max_row + 1):
                    worksheet.row_dimensions[row_num].height = 20
                    is_even = (row_num % 2 == 0)

                    for col_num in range(1, worksheet.max_column + 1):
                        cell = worksheet.cell(row=row_num, column=col_num)
                        cell.font = fonte_dados
                        cell.border = borda_fina

                        if is_even:
                            cell.fill = fill_zebra

                        nome_coluna = worksheet.cell(row=1, column=col_num).value
                        if nome_coluna in ["Telefone/WhatsApp", "Idade", "Cidade", "Escolaridade", "Status"]:
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                        else:
                            cell.alignment = Alignment(horizontal="left", vertical="center")

                        if nome_coluna == "Status":
                            if cell.value == "APROVADO":
                                cell.fill = fill_status_aprovado
                                cell.font = fonte_aprovado
                            elif cell.value == "REPROVADO":
                                cell.fill = fill_status_reprovado
                                cell.font = fonte_reprovado

                # Corrigido: o cálculo de max_len agora é efetivamente aplicado
                # à largura da coluna (antes era calculado e descartado).
                for col in worksheet.columns:
                    max_len = max(len(str(cell.value or '')) for cell in col)
                    coluna_letra = col[0].column_letter
                    worksheet.column_dimensions[coluna_letra].width = min(max_len + 4, 45)

        print(f"✅ Planilha '{nome_planilha}' gerada com {len(df_geral)} currículo(s) processado(s).")
    else:
        print("Nenhum PDF encontrado na pasta 'curriculos'.")
